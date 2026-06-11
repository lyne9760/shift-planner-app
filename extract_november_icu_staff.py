from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUTPUT = Path(__file__).with_name("2025年11月心胸外科ICU一线人员提取.xlsx")

STAFF = [
    ("强之玮", "规培", 2024, "三年", "耳鼻咽喉科", "并轨研究生", "无证", "24217298", "可副班", ""),
    ("严晓棠", "规培", 2022, "一年", "麻醉科", "外院轮转人员", "无证", "w2024231", "可副班", ""),
    ("黄沐晨", "规培", 2024, "三年", "耳鼻咽喉科", "并轨研究生", "无证", "24217295", "可副班", ""),
    ("卿时珍", "规培", 2023, "三年", "外科(胸心外科方向)", "并轨研究生", "有证", "23217212", "主班候选", "高年资胸心外科：三年培训且有证"),
    ("王娟", "规培", 2024, "三年", "耳鼻咽喉科", "并轨研究生", "无证", "24217300", "可副班", ""),
    ("王婷艳", "规培", 2024, "三年", "耳鼻咽喉科", "并轨研究生", "无证", "24217301", "可副班", ""),
    ("彭翼", "规培", 2023, "三年", "重症医学科", "并轨研究生", "有证", "23217045", "主班候选", "重症医学科"),
    ("苏晓群", "规培", 2024, "三年", "耳鼻咽喉科", "并轨研究生", "无证", "24217299", "可副班", ""),
    ("卢泽广", "规培", 2022, "一年", "麻醉科", "外院轮转人员", "无证", "w2024224", "可副班", ""),
    ("廖剑波", "规培", 2023, "三年", "外科(胸心外科方向)", "并轨研究生", "有证", "23217210", "主班候选", "高年资胸心外科：三年培训且有证"),
    ("黄嘉梁", "规培", 2023, "三年", "麻醉科", "并轨研究生", "有证", "23217097", "主班候选", "高年资麻醉：三年培训且有证"),
    ("张卫文", "规培", 2022, "一年", "麻醉科", "外院轮转人员", "无证", "w2024267", "主班候选", "用户指定与谭天互换"),
    ("谭天", "规培", "", "", "麻醉科", "", "", "", "可副班", "原表仅提供姓名、麻醉科及电话"),
    ("梁炜琳", "进修", "", "", "", "", "", "", "主班候选", "进修"),
    ("陈锵", "进修", "", "", "", "", "", "", "主班候选", "进修"),
    ("孟宽", "进修", "", "", "", "", "", "", "主班候选", "进修"),
]

EXCLUDED = [
    ("吴艺佳", "黄埔", "不参与", "用户确认不参加一线值班"),
    ("王斯琦", "黄埔", "不参与", "原表明确标注“不值班”"),
]


def build_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "一线人员"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    headers = [
        "姓名",
        "来源",
        "届别",
        "培训年限",
        "培训专业",
        "人员类型",
        "处方权",
        "工号",
        "岗位判断",
        "备注",
    ]
    header_fill = PatternFill("solid", fgColor="9FC5E8")
    primary_fill = PatternFill("solid", fgColor="D9EAD3")
    pending_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row, values in enumerate(STAFF, 2):
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row, column, value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        if values[8] == "主班候选":
            for cell in sheet[row]:
                cell.fill = primary_fill
        elif values[8] == "待确认":
            for cell in sheet[row]:
                cell.fill = pending_fill

    widths = [12, 10, 10, 12, 24, 22, 10, 14, 14, 35]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[chr(64 + index)].width = width

    excluded = workbook.create_sheet("不参与人员")
    excluded.append(["姓名", "来源", "状态", "原因"])
    for row in EXCLUDED:
        excluded.append(row)
    for cell in excluded[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    for row in excluded.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
    excluded.column_dimensions["A"].width = 12
    excluded.column_dimensions["B"].width = 12
    excluded.column_dimensions["C"].width = 12
    excluded.column_dimensions["D"].width = 35

    notes = workbook.create_sheet("提取说明")
    notes.append(["项目", "结果"])
    notes.append(["可参与一线人数", 16])
    notes.append(["规培", 13])
    notes.append(["进修", 3])
    notes.append(["黄埔", 0])
    notes.append(["主班候选（8人）", "彭翼、黄嘉梁、卿时珍、廖剑波、张卫文、梁炜琳、陈锵、孟宽"])
    notes.append(["副班候选（8人）", "强之玮、严晓棠、黄沐晨、王娟、王婷艳、苏晓群、卢泽广、谭天"])
    notes.append(["扩充原则", "优先进修、重症、高年资麻醉及高年资胸心外科，使主副班人数达到8比8"])
    notes.append(["排除", "吴艺佳：用户确认不参加一线值班；王斯琦：原表明确标注不值班"])
    for cell in notes[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
    notes.column_dimensions["A"].width = 28
    notes.column_dimensions["B"].width = 60
    for row in notes.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    return workbook


if __name__ == "__main__":
    build_workbook().save(OUTPUT)
    print(OUTPUT)
