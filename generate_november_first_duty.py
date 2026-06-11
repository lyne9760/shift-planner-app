from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path(__file__).with_name("2025年11月一线排班试排.xlsx")

GROUPS = {
    "A组": ["彭翼", "梁炜琳", "张卫文", "严晓棠", "强之玮", "王娟"],
    "B组": ["卿时珍", "黄嘉梁", "陈锵", "吴艺佳", "卢泽广", "王婷艳"],
    "C组": ["廖剑波", "孟宽", "王斯琦", "谭天", "黄沐晨", "苏晓群"],
}

MAIN = ["彭翼", "梁炜琳", "张卫文", "卿时珍", "黄嘉梁", "陈锵", "廖剑波", "孟宽"]
SECONDARY = ["严晓棠", "强之玮", "王娟", "卢泽广", "王婷艳", "谭天", "黄沐晨", "苏晓群"]
BINDINGS = {
    "彭翼": "王婷艳",
    "梁炜琳": "谭天",
    "张卫文": "苏晓群",
    "卿时珍": "严晓棠",
    "黄嘉梁": "黄沐晨",
    "陈锵": "强之玮",
    "廖剑波": "卢泽广",
    "孟宽": "王娟",
}
NON_DUTY_MAIN_TYPE = ["吴艺佳", "王斯琦"]
TRAINEES = {"梁炜琳", "陈锵", "孟宽"}
OCTOBER_ICU = {
    "严晓棠", "卢泽广", "孟宽", "彭翼", "梁炜琳",
    "王婷艳", "苏晓群", "陈锵", "黄嘉梁", "谭天",
}
GROUP_OF = {person: group for group, people in GROUPS.items() for person in people}
DATES = [date(2025, 11, day) for day in range(1, 31)]


def quality(duty_date):
    if duty_date.weekday() in (4, 5):
        return "无补休"
    if duty_date.weekday() == 3:
        return "周四黄金班"
    if duty_date.weekday() == 6:
        return "周日"
    return "可补休"


def random_role_sequence(people):
    """Build four rounds, so repeated duties naturally stay well separated."""
    sequence = []
    last_position = {}
    starts = (0, 8, 16, 24)
    lengths = (8, 8, 8, 6)
    for start, length in zip(starts, lengths):
        for _ in range(500):
            block = random.sample(people, length)
            if all(
                person not in last_position
                or start + index - last_position[person] >= 3
                for index, person in enumerate(block)
            ):
                break
        else:
            return None
        for index, person in enumerate(block):
            last_position[person] = start + index
        sequence.extend(block)
    return sequence


def secondary_sequence(main_sequence):
    """Backtrack by day to satisfy cross-group, quota and interval constraints."""
    target_four = set(random.sample(SECONDARY, 6))
    quotas = {person: 4 if person in target_four else 3 for person in SECONDARY}
    sequence = [None] * 30
    last = {person: -99 for person in SECONDARY}

    def search(day_index):
        if day_index == 30:
            return all(value == 0 for value in quotas.values())
        remaining_days = 30 - day_index
        candidates = [
            person
            for person in SECONDARY
            if quotas[person] > 0
            and day_index - last[person] >= 3
            and GROUP_OF[person] != GROUP_OF[main_sequence[day_index]]
        ]
        if not candidates:
            return False
        random.shuffle(candidates)
        candidates.sort(
            key=lambda person: (
                -quotas[person],
                last[person],
            )
        )
        for person in candidates:
            old_last = last[person]
            sequence[day_index] = person
            quotas[person] -= 1
            last[person] = day_index
            if all(value <= (remaining_days - 1 + 2) // 3 + 1 for value in quotas.values()):
                if search(day_index + 1):
                    return True
            quotas[person] += 1
            last[person] = old_last
            sequence[day_index] = None
        return False

    return sequence if search(0) else None


def schedule_score(main_sequence, secondary):
    score = 0
    for people, sequence in ((MAIN, main_sequence), (SECONDARY, secondary)):
        counts = Counter(sequence)
        category_counts = {
            person: Counter(
                quality(duty_date)
                for duty_date, assigned in zip(DATES, sequence)
                if assigned == person
            )
            for person in people
        }
        for category in ("无补休", "周四黄金班", "周日", "可补休"):
            values = [category_counts[person][category] for person in people]
            score += (max(values) - min(values)) ** 2 * 12
        for person in people:
            premium = (
                category_counts[person]["周日"]
                + 2.5 * category_counts[person]["周四黄金班"]
            )
            bad = category_counts[person]["无补休"]
            if counts[person] == 4:
                score += bad * 70 - premium * 55
                if premium == 0:
                    score += 120
            else:
                score += premium * 100 - bad * 35

    pair_counts = Counter(zip(main_sequence, secondary))
    score += sum((count - 1) ** 2 * 30 for count in pair_counts.values() if count > 1)
    group_pairs = Counter(
        (GROUP_OF[main], GROUP_OF[sub])
        for main, sub in zip(main_sequence, secondary)
    )
    if len(group_pairs) < 6:
        score += 1000
    else:
        score += (max(group_pairs.values()) - min(group_pairs.values())) ** 2 * 10

    for index in (0, 1):
        if (
            main_sequence[index] not in OCTOBER_ICU
            and secondary[index] not in OCTOBER_ICU
        ):
            score += 10000
    return score


def generate_schedule():
    best = None
    random.seed(202511)
    for _ in range(120000):
        main_sequence = random_role_sequence(MAIN)
        if not main_sequence:
            continue
        secondary = [BINDINGS[main] for main in main_sequence]
        score = schedule_score(main_sequence, secondary)
        if best is None or score < best[0]:
            best = (score, main_sequence[:], secondary[:])
    if best is None:
        raise RuntimeError("No valid schedule found")
    return list(zip(DATES, best[1], best[2]))


def add_group_sheet(workbook):
    sheet = workbook.create_sheet("三组名单")
    headers = ["小组", "姓名", "类型", "是否排班", "固定搭档", "说明"]
    for column, header in enumerate(headers, 1):
        sheet.cell(1, column, header)
    main_set = set(MAIN) | set(NON_DUTY_MAIN_TYPE)
    row = 2
    for group, people in GROUPS.items():
        for person in people:
            role = "主班类型" if person in main_set else "副班类型"
            active = "否" if person in NON_DUTY_MAIN_TYPE else "是"
            note = "进修" if person in TRAINEES else ""
            partner = ""
            if person in BINDINGS:
                partner = BINDINGS[person]
            elif person in BINDINGS.values():
                partner = next(main for main, sub in BINDINGS.items() if sub == person)
            if person in NON_DUTY_MAIN_TYPE:
                note = "仅参与分组平衡，不排一线"
            sheet.append([group, person, role, active, partner, note])
            row += 1
    return sheet


def build_workbook(schedule):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "一线排班"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    headers = ["日期", "星期", "主班", "主班组", "副班", "副班组", "班次属性", "前月ICU保障"]
    header_fill = PatternFill("solid", fgColor="9FC5E8")
    weekend_fill = PatternFill("solid", fgColor="666666")
    golden_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    weekday_names = "一二三四五六日"

    for column, header in enumerate(headers, 1):
        cell = sheet.cell(1, column, header)
        cell.fill = header_fill
        cell.font = Font(bold=True)

    for row, (duty_date, main, secondary) in enumerate(schedule, 2):
        guarantee = ""
        if duty_date.day <= 2:
            old_people = [
                person for person in (main, secondary) if person in OCTOBER_ICU
            ]
            guarantee = "、".join(old_people)
        values = [
            duty_date,
            f"周{weekday_names[duty_date.weekday()]}",
            main,
            GROUP_OF[main],
            secondary,
            GROUP_OF[secondary],
            quality(duty_date),
            guarantee,
        ]
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row, column, value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row, 1).number_format = "m月d日"
        if duty_date.weekday() in (5, 6):
            for cell in sheet[row]:
                cell.fill = weekend_fill
                cell.font = Font(color="FFFFFF")
        elif duty_date.weekday() == 3:
            for cell in sheet[row]:
                cell.fill = golden_fill

    widths = [13, 10, 12, 10, 12, 10, 15, 24]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    group_sheet = add_group_sheet(workbook)
    stats = workbook.create_sheet("统计与校验")
    stats_headers = ["岗位", "姓名", "组别", "总次数", "无补休(五六)", "周日", "周四黄金班", "其他可补休", "值班日期"]
    stats.append(stats_headers)
    for role, people, position in (("主班", MAIN, 1), ("副班", SECONDARY, 2)):
        for person in people:
            duties = [item[0] for item in schedule if item[position] == person]
            categories = Counter(quality(duty_date) for duty_date in duties)
            stats.append([
                role,
                person,
                GROUP_OF[person],
                len(duties),
                categories["无补休"],
                categories["周日"],
                categories["周四黄金班"],
                categories["可补休"],
                "、".join(f"11/{duty_date.day}" for duty_date in duties),
            ])

    checks = [
        ("每天主副班完整", len(schedule) == 30),
        ("所有主副班均跨组", all(GROUP_OF[m] != GROUP_OF[s] for _, m, s in schedule)),
        ("所有班次符合固定搭档", all(BINDINGS[m] == s for _, m, s in schedule)),
        ("吴艺佳和王斯琦未排班", all(p not in (m, s) for _, m, s in schedule for p in NON_DUTY_MAIN_TYPE)),
        ("11月1日前月ICU人员保障", any(p in OCTOBER_ICU for p in schedule[0][1:])),
        ("11月2日前月ICU人员保障", any(p in OCTOBER_ICU for p in schedule[1][1:])),
    ]
    start_row = stats.max_row + 2
    for offset, (name, passed) in enumerate(checks):
        stats.cell(start_row + offset, 1, name)
        stats.cell(start_row + offset, 2, "通过" if passed else "未通过")

    for current_sheet in (group_sheet, stats):
        current_sheet.freeze_panes = "A2"
        current_sheet.sheet_view.showGridLines = False
        for cell in current_sheet[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
        for row in current_sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in range(1, current_sheet.max_column + 1):
            current_sheet.column_dimensions[get_column_letter(column)].width = 16
    stats.column_dimensions["I"].width = 45
    return workbook


if __name__ == "__main__":
    result = generate_schedule()
    build_workbook(result).save(OUTPUT)
    print(OUTPUT)
