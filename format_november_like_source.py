from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(__file__).parent
SOURCE = BASE / "2025年11月一线二线合并排班表.xlsx"
OUTPUT = BASE / "2025年11月一线二线值班表_横版.xlsx"


def read_combined():
    workbook = load_workbook(SOURCE, data_only=True)
    sheet = workbook["一线二线总表"]
    rows = {}
    for row in range(2, 32):
        raw_date = sheet.cell(row, 1).value
        duty_date = raw_date.date() if isinstance(raw_date, datetime) else raw_date
        rows[duty_date] = {
            "main": sheet.cell(row, 3).value,
            "main_group": sheet.cell(row, 4).value,
            "secondary": sheet.cell(row, 5).value,
            "secondary_group": sheet.cell(row, 6).value,
            "backup": sheet.cell(row, 7).value,
            "backup_status": sheet.cell(row, 8).value,
        }
    return rows


def build_workbook():
    schedule = read_combined()
    dates = [date(2025, 11, day) for day in range(1, 31)]
    pairs = []
    seen = set()
    for duty_date in dates:
        item = schedule[duty_date]
        pair = (
            item["main"],
            item["main_group"],
            item["secondary"],
            item["secondary_group"],
        )
        if pair not in seen:
            seen.add(pair)
            pairs.append(pair)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.freeze_panes = "C2"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    thin = Side(style="thin", color="666666")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    month_fill = PatternFill("solid", fgColor="D9EAF7")
    weekend_fill = PatternFill("solid", fgColor="666666")
    main_fill = PatternFill("solid", fgColor="D9EAD3")
    secondary_fill = PatternFill("solid", fgColor="D9EAF7")
    backup_fill = PatternFill("solid", fgColor="EADCF8")
    covered_fill = PatternFill("solid", fgColor="F4CCCC")

    sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=32)
    month_cell = sheet.cell(1, 3, "2025年11月")
    month_cell.fill = month_fill
    month_cell.font = Font(size=14, bold=True)
    month_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.cell(2, 1, "一值")
    sheet.cell(2, 2, "固定搭档（主班 / 副班）")
    for column, duty_date in enumerate(dates, 3):
        cell = sheet.cell(2, column, duty_date.day)
        if duty_date.weekday() >= 5:
            cell.fill = weekend_fill
            cell.font = Font(color="FFFFFF", bold=True)

    first_start = 3
    for offset, (main, main_group, secondary, secondary_group) in enumerate(pairs):
        row = first_start + offset
        sheet.cell(row, 2, f"{main}（{main_group}） / {secondary}（{secondary_group}）")
        sheet.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center")
        for column, duty_date in enumerate(dates, 3):
            item = schedule[duty_date]
            cell = sheet.cell(row, column)
            if item["main"] == main and item["secondary"] == secondary:
                cell.value = "值"
                cell.fill = main_fill
                cell.font = Font(bold=True, color="006100")
            elif duty_date.weekday() >= 5:
                cell.fill = weekend_fill

    sheet.merge_cells(
        start_row=first_start,
        start_column=1,
        end_row=first_start + len(pairs) - 1,
        end_column=1,
    )
    sheet.cell(first_start, 1, "一值")

    backup_header = first_start + len(pairs)
    sheet.cell(backup_header, 1, "二线")
    sheet.cell(backup_header, 2, "人员")
    for column, duty_date in enumerate(dates, 3):
        cell = sheet.cell(backup_header, column, duty_date.day)
        if duty_date.weekday() >= 5:
            cell.fill = weekend_fill
            cell.font = Font(color="FFFFFF", bold=True)

    backup_people = "ABCDE"
    backup_start = backup_header + 1
    for offset, person in enumerate(backup_people):
        row = backup_start + offset
        sheet.cell(row, 2, person)
        for column, duty_date in enumerate(dates, 3):
            cell = sheet.cell(row, column)
            item = schedule[duty_date]
            if item["backup"] == person:
                cell.value = "值"
                cell.fill = backup_fill
                cell.font = Font(bold=True, color="5B2C83")
            elif duty_date.weekday() >= 5:
                cell.fill = weekend_fill

    sheet.merge_cells(
        start_row=backup_start,
        start_column=1,
        end_row=backup_start + len(backup_people) - 1,
        end_column=1,
    )
    sheet.cell(backup_start, 1, "二线")

    covered_row = backup_start + len(backup_people)
    sheet.cell(covered_row, 1, "二值覆盖")
    sheet.cell(covered_row, 2, "无需另排二线")
    for column, duty_date in enumerate(dates, 3):
        item = schedule[duty_date]
        cell = sheet.cell(covered_row, column)
        if item["backup"] == "无需另排":
            cell.value = item["backup_status"].replace("（二值）", "")
            cell.fill = covered_fill
            cell.alignment = Alignment(
                text_rotation=90, horizontal="center", vertical="center"
            )
        elif duty_date.weekday() >= 5:
            cell.fill = weekend_fill
    sheet.row_dimensions[covered_row].height = 62

    weekday_row = covered_row + 1
    sheet.cell(weekday_row, 1, "星期")
    weekday_names = "一二三四五六日"
    for column, duty_date in enumerate(dates, 3):
        cell = sheet.cell(
            weekday_row, column, f"周{weekday_names[duty_date.weekday()]}"
        )
        if duty_date.weekday() >= 5:
            cell.fill = weekend_fill
            cell.font = Font(color="FFFFFF")

    for row in sheet.iter_rows(
        min_row=1, max_row=weekday_row, min_col=1, max_col=32
    ):
        for cell in row:
            cell.border = border
            if cell.row != covered_row or cell.column <= 2:
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

    sheet.column_dimensions["A"].width = 11.5
    sheet.column_dimensions["B"].width = 30
    for column in range(3, 33):
        sheet.column_dimensions[get_column_letter(column)].width = 3.6
    sheet.print_area = f"A1:AF{weekday_row}"

    notes = workbook.create_sheet("说明")
    notes.append(["项目", "说明"])
    notes.append(["一值", "每行是一组固定主班/副班搭档；值班日期标“值”"])
    notes.append(["二线", "A-E五人值班日期标“值”"])
    notes.append(["二值覆盖", "张宝、李思、成艳美值二值的日期，不另排二线"])
    notes.append(["深色日期", "沿用原值班表样式，表示周六、周日"])
    notes.column_dimensions["A"].width = 18
    notes.column_dimensions["B"].width = 60
    for row in notes.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for cell in notes[1]:
        cell.fill = month_fill
        cell.font = Font(bold=True)
    return workbook


if __name__ == "__main__":
    build_workbook().save(OUTPUT)
    print(OUTPUT)
