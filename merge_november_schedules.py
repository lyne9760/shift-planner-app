from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(__file__).parent
FIRST_DUTY_FILE = BASE / "2025年11月一线排班试排.xlsx"
BACKUP_FILE = BASE / "二线排班试排_ABCDE_2025年11月.xlsx"
OUTPUT = BASE / "2025年11月一线二线合并排班表.xlsx"


def read_first_duty():
    workbook = load_workbook(FIRST_DUTY_FILE, data_only=True)
    sheet = workbook["一线排班"]
    result = {}
    for row in range(2, sheet.max_row + 1):
        raw_date = sheet.cell(row, 1).value
        if isinstance(raw_date, datetime):
            duty_date = raw_date.date()
        elif isinstance(raw_date, date):
            duty_date = raw_date
        else:
            continue
        result[duty_date] = {
            "weekday": sheet.cell(row, 2).value,
            "main": sheet.cell(row, 3).value,
            "main_group": sheet.cell(row, 4).value,
            "secondary": sheet.cell(row, 5).value,
            "secondary_group": sheet.cell(row, 6).value,
            "quality": sheet.cell(row, 7).value,
            "previous_month": sheet.cell(row, 8).value,
        }
    return result, workbook


def read_backup():
    workbook = load_workbook(BACKUP_FILE, data_only=True)
    sheet = workbook["二线排班"]
    result = {}
    for column in range(3, sheet.max_column + 1):
        day = sheet.cell(2, column).value
        if not isinstance(day, int):
            continue
        duty_date = date(2025, 11, day)
        assigned = None
        for row in range(4, 9):
            if sheet.cell(row, column).value == "值":
                assigned = sheet.cell(row, 2).value
                break
        covered = sheet.cell(9, column).value
        result[duty_date] = {
            "backup": assigned,
            "covered": covered,
        }
    return result, workbook


def copy_sheet_values(source, target):
    for row in source.iter_rows():
        for cell in row:
            new_cell = target[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)
    for key, dimension in source.column_dimensions.items():
        target.column_dimensions[key].width = dimension.width
        target.column_dimensions[key].hidden = dimension.hidden
    for key, dimension in source.row_dimensions.items():
        target.row_dimensions[key].height = dimension.height
        target.row_dimensions[key].hidden = dimension.hidden
    for merged_range in source.merged_cells.ranges:
        target.merge_cells(str(merged_range))
    target.freeze_panes = source.freeze_panes
    target.sheet_view.showGridLines = source.sheet_view.showGridLines


def build_workbook():
    first_duty, first_workbook = read_first_duty()
    backup, backup_workbook = read_backup()

    workbook = Workbook()
    combined = workbook.active
    combined.title = "一线二线总表"
    combined.freeze_panes = "A2"
    combined.sheet_view.showGridLines = False

    headers = [
        "日期",
        "星期",
        "主班",
        "主班组",
        "副班",
        "副班组",
        "二线",
        "二线状态",
        "班次属性",
        "前月ICU保障",
    ]
    header_fill = PatternFill("solid", fgColor="9FC5E8")
    weekend_fill = PatternFill("solid", fgColor="666666")
    golden_fill = PatternFill("solid", fgColor="FFF2CC")
    covered_fill = PatternFill("solid", fgColor="F4CCCC")
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for column, header in enumerate(headers, 1):
        cell = combined.cell(1, column, header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row, day in enumerate(range(1, 31), 2):
        duty_date = date(2025, 11, day)
        first = first_duty[duty_date]
        backup_info = backup[duty_date]
        backup_name = backup_info["backup"] or "无需另排"
        backup_status = (
            backup_info["covered"]
            if backup_info["covered"]
            else "ABCDE二线"
        )
        values = [
            duty_date,
            first["weekday"],
            first["main"],
            first["main_group"],
            first["secondary"],
            first["secondary_group"],
            backup_name,
            backup_status,
            first["quality"],
            first["previous_month"],
        ]
        for column, value in enumerate(values, 1):
            cell = combined.cell(row, column, value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        combined.cell(row, 1).number_format = "m月d日"

        if backup_info["covered"]:
            combined.cell(row, 7).fill = covered_fill
            combined.cell(row, 8).fill = covered_fill
        if duty_date.weekday() in (5, 6):
            for column in (1, 2, 3, 4, 5, 6, 9, 10):
                combined.cell(row, column).fill = weekend_fill
                combined.cell(row, column).font = Font(color="FFFFFF")
            if not backup_info["covered"]:
                for column in (7, 8):
                    combined.cell(row, column).fill = weekend_fill
                    combined.cell(row, column).font = Font(color="FFFFFF")
        elif duty_date.weekday() == 3:
            for column in range(1, 11):
                if not (backup_info["covered"] and column in (7, 8)):
                    combined.cell(row, column).fill = golden_fill

    widths = [13, 10, 12, 10, 12, 10, 12, 20, 15, 24]
    for index, width in enumerate(widths, 1):
        combined.column_dimensions[get_column_letter(index)].width = width

    source_sheets = [
        (first_workbook["三组名单"], "三组名单"),
        (first_workbook["统计与校验"], "一线统计校验"),
        (backup_workbook["二线排班"], "二线横表"),
        (backup_workbook["统计与校验"], "二线统计校验"),
    ]
    for source, title in source_sheets:
        target = workbook.create_sheet(title)
        copy_sheet_values(source, target)
    return workbook


if __name__ == "__main__":
    build_workbook().save(OUTPUT)
    print(OUTPUT)
