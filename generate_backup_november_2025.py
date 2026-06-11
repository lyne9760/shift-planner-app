from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path(__file__).with_name("二线排班试排_ABCDE_2025年11月.xlsx")

ASSIGNMENTS = {
    date(2025, 11, 2): "B",
    date(2025, 11, 3): "D",
    date(2025, 11, 4): "E",
    date(2025, 11, 6): "A",
    date(2025, 11, 7): "B",
    date(2025, 11, 9): "D",
    date(2025, 11, 10): "C",
    date(2025, 11, 11): "E",
    date(2025, 11, 12): "B",
    date(2025, 11, 14): "A",
    date(2025, 11, 15): "C",
    date(2025, 11, 16): "E",
    date(2025, 11, 18): "B",
    date(2025, 11, 19): "D",
    date(2025, 11, 21): "E",
    date(2025, 11, 22): "A",
    date(2025, 11, 23): "D",
    date(2025, 11, 24): "C",
    date(2025, 11, 26): "A",
    date(2025, 11, 27): "E",
    date(2025, 11, 28): "D",
    date(2025, 11, 30): "C",
}

COVERED_BY_SECOND_DUTY = {
    date(2025, 11, 1): "张宝（二值）",
    date(2025, 11, 5): "成艳美（二值）",
    date(2025, 11, 8): "李思（二值）",
    date(2025, 11, 13): "张宝（二值）",
    date(2025, 11, 17): "成艳美（二值）",
    date(2025, 11, 20): "李思（二值）",
    date(2025, 11, 25): "张宝（二值）",
    date(2025, 11, 29): "成艳美（二值）",
}

# These are the dark columns in the source second-duty workbook.
HOLIDAYS = {
    date(2025, 11, day)
    for day in (1, 2, 8, 9, 15, 16, 22, 23, 29, 30)
}


def month_dates():
    current = date(2025, 11, 1)
    while current <= date(2025, 11, 30):
        yield current
        current += timedelta(days=1)


def build_workbook():
    dates = list(month_dates())
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "二线排班"
    schedule.freeze_panes = "C4"
    schedule.sheet_view.showGridLines = False

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="9FC5E8")
    holiday_fill = PatternFill("solid", fgColor="666666")
    duty_fill = PatternFill("solid", fgColor="D9EAD3")
    covered_fill = PatternFill("solid", fgColor="F4CCCC")

    schedule.merge_cells(
        start_row=1, start_column=3, end_row=1, end_column=2 + len(dates)
    )
    title = schedule.cell(1, 3, "2025年11月二线排班试排")
    title.font = Font(size=14, bold=True)
    title.alignment = Alignment(horizontal="center")

    schedule.cell(2, 1, "岗位")
    schedule.cell(2, 2, "人员")
    weekday_names = "一二三四五六日"
    for column, duty_date in enumerate(dates, 3):
        date_cell = schedule.cell(2, column, duty_date.day)
        weekday_cell = schedule.cell(
            3, column, f"周{weekday_names[duty_date.weekday()]}"
        )
        if duty_date in HOLIDAYS:
            for cell in (date_cell, weekday_cell):
                cell.fill = holiday_fill
                cell.font = Font(color="FFFFFF", bold=True)

    for row, person in enumerate("ABCDE", 4):
        schedule.cell(row, 1, "二线")
        schedule.cell(row, 2, person)
        for column, duty_date in enumerate(dates, 3):
            cell = schedule.cell(row, column)
            if ASSIGNMENTS.get(duty_date) == person:
                cell.value = "值"
                if duty_date in HOLIDAYS:
                    cell.fill = holiday_fill
                    cell.font = Font(bold=True, color="FFFFFF")
                else:
                    cell.fill = duty_fill
                    cell.font = Font(bold=True, color="006100")
            elif duty_date in COVERED_BY_SECOND_DUTY:
                cell.value = "—"
                cell.fill = covered_fill
            elif duty_date in HOLIDAYS:
                cell.fill = holiday_fill

    note_row = 9
    schedule.merge_cells(
        start_row=note_row, start_column=1, end_row=note_row, end_column=2
    )
    schedule.cell(note_row, 1, "已有二值，无需另排")
    for column, duty_date in enumerate(dates, 3):
        if duty_date in COVERED_BY_SECOND_DUTY:
            cell = schedule.cell(
                note_row, column, COVERED_BY_SECOND_DUTY[duty_date]
            )
            cell.fill = covered_fill
            cell.alignment = Alignment(
                text_rotation=90, horizontal="center", vertical="center"
            )
        elif duty_date in HOLIDAYS:
            schedule.cell(note_row, column).fill = holiday_fill
    schedule.row_dimensions[note_row].height = 85

    for row in schedule.iter_rows(
        min_row=2, max_row=note_row, min_col=1, max_col=2 + len(dates)
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
    for cell in schedule[2]:
        if cell.column < 3 or cell.value not in range(1, 31):
            cell.font = Font(bold=True)
    for cell in schedule[3]:
        if cell.fill != holiday_fill:
            cell.font = Font(size=9)

    schedule.column_dimensions["A"].width = 13
    schedule.column_dimensions["B"].width = 10
    for column in range(3, 3 + len(dates)):
        schedule.column_dimensions[get_column_letter(column)].width = 5

    stats = workbook.create_sheet("统计与校验")
    stats.sheet_view.showGridLines = False
    headers = [
        "人员",
        "总次数",
        "周一至周四",
        "周五",
        "周六",
        "周日",
        "深色节假日",
        "最小间隔（天）",
        "值班日期",
    ]
    for column, header in enumerate(headers, 1):
        cell = stats.cell(1, column, header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row, person in enumerate("ABCDE", 2):
        duty_dates = sorted(
            duty_date
            for duty_date, assigned in ASSIGNMENTS.items()
            if assigned == person
        )
        gaps = [
            (later - earlier).days
            for earlier, later in zip(duty_dates, duty_dates[1:])
        ]
        values = [
            person,
            len(duty_dates),
            sum(duty_date.weekday() <= 3 for duty_date in duty_dates),
            sum(duty_date.weekday() == 4 for duty_date in duty_dates),
            sum(duty_date.weekday() == 5 for duty_date in duty_dates),
            sum(duty_date.weekday() == 6 for duty_date in duty_dates),
            sum(duty_date in HOLIDAYS for duty_date in duty_dates),
            min(gaps),
            "、".join(f"11/{duty_date.day}" for duty_date in duty_dates),
        ]
        for column, value in enumerate(values, 1):
            cell = stats.cell(row, column, value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    notes = [
        "李思二值：11月8日、20日；张宝二值：11月1日、13日、25日；成艳美二值：11月5日、17日、29日。以上日期不另排二线。",
        "原二值表深色日期：11月1-2日、8-9日、15-16日、22-23日、29-30日，按节假日统计。",
        "A、B、C各4次，D、E各5次；4个周五由A、B、D、E各承担1次；相邻两次值班至少间隔4天。",
    ]
    for row, note in enumerate(notes, 7):
        stats.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        stats.cell(row, 1, note)
        stats.cell(row, 1).alignment = Alignment(wrap_text=True)
        stats.cell(row, 1).font = Font(color="666666")

    widths = {
        "A": 10,
        "B": 10,
        "C": 14,
        "D": 8,
        "E": 8,
        "F": 8,
        "G": 14,
        "H": 16,
        "I": 55,
    }
    for column, width in widths.items():
        stats.column_dimensions[column].width = width
    return workbook


if __name__ == "__main__":
    build_workbook().save(OUTPUT)
    print(OUTPUT)
