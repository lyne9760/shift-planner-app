from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


OUTPUT = Path(__file__).with_name("二线排班试排_ABCD_2025-05-06至06-03.xlsx")

ASSIGNMENTS = {
    date(2025, 5, 6): "D",
    date(2025, 5, 7): "C",
    date(2025, 5, 8): "B",
    date(2025, 5, 9): "A",
    date(2025, 5, 11): "D",
    date(2025, 5, 12): "A",
    date(2025, 5, 13): "C",
    date(2025, 5, 14): "B",
    date(2025, 5, 16): "D",
    date(2025, 5, 17): "A",
    date(2025, 5, 18): "C",
    date(2025, 5, 19): "B",
    date(2025, 5, 20): "D",
    date(2025, 5, 22): "A",
    date(2025, 5, 23): "C",
    date(2025, 5, 24): "B",
    date(2025, 5, 26): "D",
    date(2025, 5, 27): "A",
    date(2025, 5, 29): "C",
    date(2025, 5, 30): "B",
    date(2025, 5, 31): "D",
    date(2025, 6, 1): "A",
    date(2025, 6, 2): "C",
    date(2025, 6, 3): "B",
}

COVERED_BY_SECOND_DUTY = {
    date(2025, 5, 10): "刘芙蓉（二值）",
    date(2025, 5, 15): "李思（二值）",
    date(2025, 5, 21): "刘芙蓉（二值）",
    date(2025, 5, 25): "李思（二值）",
    date(2025, 5, 28): "刘芙蓉（二值）",
}


def date_range(start, end):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def build_workbook():
    dates = list(date_range(date(2025, 5, 6), date(2025, 6, 3)))
    workbook = Workbook()
    schedule = workbook.active
    schedule.title = "二线排班"
    schedule.freeze_panes = "C4"
    schedule.sheet_view.showGridLines = False

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    weekend_fill = PatternFill("solid", fgColor="FFF2CC")
    covered_fill = PatternFill("solid", fgColor="F4CCCC")
    duty_fill = PatternFill("solid", fgColor="D9EAD3")
    header_fill = PatternFill("solid", fgColor="9FC5E8")

    schedule.merge_cells(
        start_row=1, start_column=3, end_row=1, end_column=2 + len(dates)
    )
    title = schedule.cell(1, 3, "二线排班试排（2025年5月6日—6月3日）")
    title.font = Font(size=14, bold=True)
    title.alignment = Alignment(horizontal="center")

    schedule.cell(2, 1, "岗位")
    schedule.cell(2, 2, "人员")
    weekday_names = "一二三四五六日"
    for column, duty_date in enumerate(dates, 3):
        schedule.cell(2, column, duty_date.day)
        schedule.cell(3, column, f"周{weekday_names[duty_date.weekday()]}")
        if duty_date.weekday() >= 4:
            schedule.cell(2, column).fill = weekend_fill
            schedule.cell(3, column).fill = weekend_fill

    for row, person in enumerate("ABCD", 4):
        schedule.cell(row, 1, "二线")
        schedule.cell(row, 2, person)
        for column, duty_date in enumerate(dates, 3):
            cell = schedule.cell(row, column)
            if ASSIGNMENTS.get(duty_date) == person:
                cell.value = "值"
                cell.fill = duty_fill
                cell.font = Font(bold=True, color="006100")
            elif duty_date in COVERED_BY_SECOND_DUTY:
                cell.value = "—"
                cell.fill = covered_fill

    note_row = 8
    schedule.cell(note_row, 1, "无需另排二线")
    schedule.merge_cells(
        start_row=note_row, start_column=1, end_row=note_row, end_column=2
    )
    for column, duty_date in enumerate(dates, 3):
        if duty_date in COVERED_BY_SECOND_DUTY:
            cell = schedule.cell(
                note_row, column, COVERED_BY_SECOND_DUTY[duty_date]
            )
            cell.fill = covered_fill
            cell.alignment = Alignment(
                text_rotation=90, horizontal="center", vertical="center"
            )
    schedule.row_dimensions[note_row].height = 85

    for row in schedule.iter_rows(
        min_row=2, max_row=note_row, min_col=1, max_col=2 + len(dates)
    ):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
    for cell in schedule[2]:
        cell.font = Font(bold=True)
    for cell in schedule[3]:
        cell.font = Font(size=9)
    schedule.column_dimensions["A"].width = 12
    schedule.column_dimensions["B"].width = 10
    for column in range(3, 3 + len(dates)):
        schedule.column_dimensions[get_column_letter(column)].width = 5

    stats = workbook.create_sheet("统计与校验")
    stats.sheet_view.showGridLines = False
    headers = [
        "人员",
        "总次数",
        "周一至周四",
        "周五",
        "周六",
        "周日",
        "最小间隔（天）",
        "值班日期",
    ]
    for column, header in enumerate(headers, 1):
        cell = stats.cell(1, column, header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    for row, person in enumerate("ABCD", 2):
        duty_dates = sorted(
            duty_date
            for duty_date, assigned in ASSIGNMENTS.items()
            if assigned == person
        )
        categories = [
            sum(duty_date.weekday() <= 3 for duty_date in duty_dates),
            sum(duty_date.weekday() == 4 for duty_date in duty_dates),
            sum(duty_date.weekday() == 5 for duty_date in duty_dates),
            sum(duty_date.weekday() == 6 for duty_date in duty_dates),
        ]
        gaps = [
            (later - earlier).days
            for earlier, later in zip(duty_dates, duty_dates[1:])
        ]
        values = [
            person,
            len(duty_dates),
            *categories,
            min(gaps),
            "、".join(f"{duty_date.month}/{duty_date.day}" for duty_date in duty_dates),
        ]
        for column, value in enumerate(values, 1):
            cell = stats.cell(row, column, value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    notes = [
        "规则：2025年5月10、15、21、25、28日已有李思或刘芙蓉值二值，因此不另排二线。",
        "A-D各6次；同一人相邻两次值班日期至少相差3天（中间至少空2天）。",
        "周五共4次，四人各1次；周六、周日各只有3个可排日期，无法做到四人各1次。",
    ]
    for row, note in enumerate(notes, 7):
        stats.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        stats.cell(row, 1, note)
        stats.cell(row, 1).alignment = Alignment(wrap_text=True)
        stats.cell(row, 1).font = Font(color="666666")

    widths = {
        "A": 10,
        "B": 10,
        "C": 14,
        "D": 8,
        "E": 8,
        "F": 8,
        "G": 16,
        "H": 55,
    }
    for column, width in widths.items():
        stats.column_dimensions[column].width = width
    return workbook


if __name__ == "__main__":
    build_workbook().save(OUTPUT)
    print(OUTPUT)
